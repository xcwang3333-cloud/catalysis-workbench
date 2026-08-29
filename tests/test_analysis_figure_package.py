from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from catalysis_workbench.application import (
    AnalysisRange,
    AnalysisSession,
    AnalysisSessionError,
    DataSeriesSpec,
    FigurePackageOptions,
    GenericXYAnalysisSpec,
    TabularMappingSpec,
    export_session_figure_package,
    source_spec_from_file,
)
from catalysis_workbench.workspace.composition import open_workspace_composition
from catalysis_workbench.workspace.evidence import open_evidence_ledger


def _mapped(path: Path, *, name: str) -> DataSeriesSpec:
    return DataSeriesSpec(
        source=source_spec_from_file(path),
        mapping=TabularMappingSpec(
            delimiter=",",
            x_column=0,
            y_column=1,
            x_role="potential",
            y_role="signal",
            x_unit="V",
            y_unit="a.u.",
            x_reference="RHE",
        ),
        display_name=name,
    )


def _saved_figure_session(
    tmp_path: Path,
) -> tuple[AnalysisSession, DataSeriesSpec, DataSeriesSpec]:
    first_path = tmp_path / "first.csv"
    second_path = tmp_path / "second.csv"
    first_path.write_text("x,y\n0,1\n1,\n2,3\n", encoding="utf-8")
    second_path.write_text("x,y\n0,3\n1,2\n2,1\n", encoding="utf-8")
    first = _mapped(first_path, name="Pb1")
    second = _mapped(second_path, name="Pb2")

    session = AnalysisSession()
    session.new_analysis("generic_xy")
    session.add_data_series_batch(((first, first_path), (second, second_path)))
    session.create_figure("processed")
    draft = session.figure_draft("processed")
    spec = (
        draft.figure_spec.updated(xlim=(0.5, 1.5))
        .with_series_style(first.data_id, label="Pb₁-N/C")
        .with_series_style(second.data_id, visible=False)
    )
    session.replace_figure_spec("processed", spec)
    session.save_project_as(tmp_path / "project")
    assert session.state.is_dirty is False
    return session, first, second


def _project_tree(root: Path) -> dict[str, bytes]:
    return {
        path.relative_to(root).as_posix(): path.read_bytes()
        for path in sorted(root.rglob("*"))
        if path.is_file() and not path.is_symlink()
    }


def test_package_exports_visible_full_science_and_updates_only_workspace_baseline(
    tmp_path: Path,
) -> None:
    session, first, second = _saved_figure_session(tmp_path)
    before = session.state
    assert before.document is not None
    before_document_sha = before.document.document_sha256

    target = tmp_path / "publication-package"
    result = export_session_figure_package(
        session,
        "processed",
        target,
        options=FigurePackageOptions(
            figure_formats=("svg",),
            source_data_formats=("xlsx", "txt"),
        ),
    )

    after = session.state
    assert target == result.package_path
    assert after.document is before.document
    assert after.document.document_sha256 == before_document_sha
    assert after.revision == before.revision
    assert after.can_undo == before.can_undo
    assert after.can_redo == before.can_redo
    assert after.is_dirty is False
    assert after.project_file_sha256 == before.project_file_sha256
    assert after.workspace_manifest_sha256 == result.workspace_manifest_sha256
    assert after.workspace_manifest_sha256 != before.workspace_manifest_sha256

    manifest = json.loads((target / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["package_sha256"] == result.package_sha256
    assert (
        manifest["figure_draft_sha256"]
        == session.figure_draft("processed").figure_sha256
    )
    assert manifest["trace_order"] == [first.data_id]
    assert set(manifest["trace_identities"]) == {first.data_id}
    assert second.data_id not in json.dumps(manifest, sort_keys=True)
    serialized = json.dumps(manifest, sort_keys=True)
    assert str(tmp_path.resolve()) not in serialized
    assert "timestamp" not in serialized.casefold()
    assert "created_at" not in serialized.casefold()

    txt = (target / "source-data" / "trace-001.txt").read_text(encoding="utf-8")
    rows = [line for line in txt.splitlines() if not line.startswith("#")]
    assert rows[0] == "x\ty\tx_missing\ty_missing"
    assert rows[1:] == [
        "0\t1\t0\t0",
        "1\tnan\t0\t1",
        "2\t3\t0\t0",
    ]
    assert not (target / "source-data" / "trace-002.txt").exists()

    workbook = load_workbook(target / "source-data.xlsx", data_only=True)
    assert workbook.sheetnames == ["Index", "Trace 001"]
    sheet = workbook["Trace 001"]
    assert sheet.max_row == 4
    assert sheet.cell(3, 1).value == 1
    assert sheet.cell(3, 2).value is None
    assert sheet.cell(3, 4).value is True

    ledger = open_evidence_ledger(before.project_root)
    assert any(record.kind == "workflow_run" for record in ledger.records)
    package_records = [
        record for record in ledger.records if record.kind == "artifact"
    ]
    assert len(package_records) == 1
    assert package_records[0].evidence_sha256 == result.manifest_sha256
    composition = open_workspace_composition(before.project_root)
    assert len(composition.figures) == 1


def test_package_semantic_identity_is_destination_independent(tmp_path: Path) -> None:
    session, _first, _second = _saved_figure_session(tmp_path)
    options = FigurePackageOptions(
        figure_formats=("svg",),
        source_data_formats=("txt",),
    )
    first = export_session_figure_package(
        session,
        "processed",
        tmp_path / "package-a",
        options=options,
    )
    second = export_session_figure_package(
        session,
        "processed",
        tmp_path / "package-b",
        options=options,
    )
    assert first.package_sha256 == second.package_sha256


def test_final_publish_failure_rolls_back_package_provenance_and_session(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, _first, _second = _saved_figure_session(tmp_path)
    before_state = session.state
    assert before_state.project_root is not None
    before_tree = _project_tree(before_state.project_root)
    target = tmp_path / "failed-package"

    import catalysis_workbench.application.analysis.export_publish as publisher

    def fail_publish(_stage: Path, _target: Path) -> None:
        raise OSError("injected final publication failure")

    monkeypatch.setattr(publisher, "_publish_stage", fail_publish)
    with pytest.raises(AnalysisSessionError, match="injected final publication failure"):
        export_session_figure_package(
            session,
            "processed",
            target,
            options=FigurePackageOptions(
                figure_formats=("svg",),
                source_data_formats=("txt",),
            ),
        )

    assert session.state == before_state
    assert not target.exists()
    assert _project_tree(before_state.project_root) == before_tree


def test_post_rename_failure_removes_only_the_exact_published_package(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, _first, _second = _saved_figure_session(tmp_path)
    before_state = session.state
    assert before_state.project_root is not None
    before_tree = _project_tree(before_state.project_root)
    target = tmp_path / "post-rename-failure"

    import catalysis_workbench.application.analysis.export_publish as publisher

    publish = publisher._publish_stage

    def move_then_fail(stage: Path, destination: Path) -> None:
        publish(stage, destination)
        raise OSError("injected post-rename failure")

    monkeypatch.setattr(publisher, "_publish_stage", move_then_fail)
    with pytest.raises(AnalysisSessionError, match="injected post-rename failure"):
        export_session_figure_package(
            session,
            "processed",
            target,
            options=FigurePackageOptions(
                figure_formats=("svg",),
                source_data_formats=("txt",),
            ),
        )

    assert session.state == before_state
    assert not target.exists()
    assert _project_tree(before_state.project_root) == before_tree


def test_post_rename_external_mutation_is_preserved_for_inspection(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, _first, _second = _saved_figure_session(tmp_path)
    before_state = session.state
    assert before_state.project_root is not None
    before_tree = _project_tree(before_state.project_root)
    target = tmp_path / "mutated-package"

    import catalysis_workbench.application.analysis.export_publish as publisher

    publish = publisher._publish_stage

    def move_mutate_then_fail(stage: Path, destination: Path) -> None:
        publish(stage, destination)
        (destination / "intruder.txt").write_text("external", encoding="utf-8")
        raise OSError("injected mutation after rename")

    monkeypatch.setattr(publisher, "_publish_stage", move_mutate_then_fail)
    with pytest.raises(
        AnalysisSessionError,
        match="destination changed",
    ):
        export_session_figure_package(
            session,
            "processed",
            target,
            options=FigurePackageOptions(
                figure_formats=("svg",),
                source_data_formats=("txt",),
            ),
        )

    assert session.state == before_state
    assert target.is_dir()
    assert (target / "intruder.txt").read_text(encoding="utf-8") == "external"
    assert _project_tree(before_state.project_root) != before_tree


def test_export_rejects_unsaved_dirty_stale_and_existing_target(tmp_path: Path) -> None:
    raw = tmp_path / "raw.csv"
    raw.write_text("x,y\n0,1\n1,2\n2,3\n", encoding="utf-8")
    spec = _mapped(raw, name="Pb")
    session = AnalysisSession()
    session.new_analysis("generic_xy")
    session.add_data_series(spec, raw)
    session.create_figure("processed")

    with pytest.raises(AnalysisSessionError, match="save the analysis project"):
        export_session_figure_package(session, "processed", tmp_path / "unsaved")

    project = tmp_path / "project"
    session.save_project_as(project)
    session.rename_analysis("dirty")
    with pytest.raises(AnalysisSessionError, match="save the current analysis project"):
        export_session_figure_package(session, "processed", tmp_path / "dirty")

    session.save_project()
    session.replace_analysis_spec(
        GenericXYAnalysisSpec(analysis_range=AnalysisRange(x_min=0.5, x_max=2.0))
    )
    session.save_project()
    assert session.figure_is_stale("processed") is True
    with pytest.raises(AnalysisSessionError, match="refresh this figure"):
        export_session_figure_package(session, "processed", tmp_path / "stale")

    session.refresh_figure("processed")
    session.save_project()
    existing = tmp_path / "existing"
    existing.mkdir()
    with pytest.raises(AnalysisSessionError, match="must not already exist"):
        export_session_figure_package(session, "processed", existing)
